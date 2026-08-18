"""Turn every source into merge-ready person records.

One adapter per source. Each returns plain dicts tagged with `source`, so
merge.build_people() can union them without knowing where anything came
from. Adapters read; they never decide -- precedence and conflicts are
merge's job.
"""

import csv
import datetime
import os
import re

import openpyxl

from . import config, flags, notes as notesmod, parse_jobtab, parse_sow
from . import parse_vcf, parse_vendors
from .normalize import clean_str, looks_like_email, looks_like_phone
from .normalize import norm_email, norm_phone
from .roles import RoleResolver


def _spec_rows():
    with open(config.source_path("rolly_spec.csv"), encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


def from_rollies(resolver=None, progress=None):
    """People tabs across the 45 rolodex workbooks.

    Reads colour as evidence (flags.classify decides whether it blocks) and
    keeps each note tagged with the workbook and tab it came from.
    """
    resolver = resolver or RoleResolver()
    out = []
    for s in _spec_rows():
        if s["kind"] != "people" or s["name_col"] == "":
            continue
        path = os.path.join(config.rollies_dir(), s["workbook"])
        # Styles are needed for the colour flags, so no read_only here.
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[s["tab"]]
        ni = int(s["name_col"])
        hdr = int(s["header_rows"] or 0)
        gi = int(s["grade_col"]) if s["grade_col"] != "" else None
        pi = int(s["phone_col"]) if s["phone_col"] != "" else None
        ei = int(s["email_col"]) if s["email_col"] != "" else None
        ri = int(s["roles_col"]) if s["roles_col"] != "" else None
        ci = int(s["city_col"]) if s["city_col"] != "" else None
        noi = int(s["notes_col"]) if s["notes_col"] != "" else None
        shortlist = s["status"] == "Short List"

        for r_i, row in enumerate(ws.iter_rows()):
            if r_i < hdr or ni >= len(row):
                continue
            raw_name = clean_str(row[ni].value)
            if not raw_name or len(raw_name) < 2:
                continue
            if looks_like_phone(raw_name) or looks_like_email(raw_name):
                continue

            def cell(i):
                return row[i] if i is not None and i < len(row) else None

            note_text = clean_str(cell(noi).value if cell(noi) else "")
            colour, _kind = flags.color_of(row[ni])
            verdict = flags.classify(note_text, colour)

            rec = {
                "source": "rolly",
                "name": raw_name,
                "phoneDigits": norm_phone(cell(pi).value) if cell(pi) and looks_like_phone(cell(pi).value) else "",
                "email": norm_email(cell(ei).value) if cell(ei) and looks_like_email(clean_str(cell(ei).value)) else "",
                "markets": [s["market"]] if s["market"] else [],
                "grade": clean_str(cell(gi).value) if cell(gi) else "",
                "shortlisted": shortlist,
                "rollyLists": ["%s › %s" % (re.sub(r"\.xlsx$", "", s["workbook"]), s["tab"])],
                "doNotHire": bool(verdict["blocking"]),
                "advisoryFlags": [f for f in verdict["flags"] if f not in verdict["blocking"]],
                "source_file": s["workbook"],
                "source_tab": s["tab"],
                "source_cell": flags.cell_ref(ni, r_i),
                "source_excerpt": note_text[:120],
            }
            # Roles a rolodex lists are CLAIMS -- the rolodex is not evidence
            # that anyone booked them. Contracts and job sheets supply worked
            # titles; merge subtracts anything proven.
            if cell(ri):
                rec["claimedSkills"] = resolver.resolve_cell(clean_str(cell(ri).value))
            if cell(ci):
                city = clean_str(cell(ci).value)
                if city and len(city) < 40:
                    rec["markets"] = rec["markets"] + [city]
            if note_text:
                rec["notes"] = ["[%s] %s" % (
                    notesmod.source_label(workbook=s["workbook"], tab=s["tab"]), t)
                    for t in notesmod.split_observations(note_text)]
            out.append(rec)
        wb.close()
        if progress:
            progress(s["workbook"], s["tab"], len(out))
    return out


def from_vcard():
    """Do-not-hire people from the phone export."""
    out = []
    for r in parse_vcf.extract_flagged(config.contacts_vcf()):
        out.append({
            "source": "vcard",
            "name": r["name"],
            "phoneDigits": r["phones"][0] if r["phones"] else "",
            "email": r["emails"][0] if r["emails"] else "",
            "doNotHire": True,
            "notes": (["[contact list] " + r["name_note"]] if r["name_note"] else []),
            "source_file": "All Contacts.vcf",
            "source_excerpt": r["raw_name"][:120],
        })
    return out


def from_vendors():
    """Home addresses from the QuickBooks export. Tax IDs never leave the
    parser, so nothing sensitive can reach a person row."""
    people, stats = parse_vendors.parse(config.vendor_csv())
    out = []
    for p in people:
        out.append({
            "source": "vendor",
            "name": p["name"], "aka": p["aka"],
            "email": p["email"], "phoneDigits": p["phoneDigits"],
            "home_base": p["home_base"], "address": p["address"],
            "state": p["state"], "zip": p["zip"],
            "source_file": "AKAV LLC_Vendor Contact List.csv",
        })
    return out, stats


def from_contracts(resolver=None):
    """Worked positions + real day rates from signed SOWs."""
    resolver = resolver or RoleResolver()
    days, _skipped, _problems = parse_sow.parse_folder(
        config.contracts_dir(), resolver)
    by_person = {}
    for d in days:
        k = d["person"].lower()
        e = by_person.setdefault(k, {
            "source": "contract", "name": d["person"], "aka": d.get("aliases", []),
            "jobTitles": [], "rates": [], "source_file": "Contracts",
        })
        for t in d["titles"]:
            if t not in e["jobTitles"]:
                e["jobTitles"].append(t)
        # Half days are excluded from rate stats.
        if d["rate"] is not None and not d["is_half"]:
            e["rates"].append(d["rate"])
    out = []
    for e in by_person.values():
        rates = sorted(e.pop("rates"))
        if rates:
            n = len(rates)
            e["dayRate"] = rates[n // 2] if n % 2 else (rates[n // 2 - 1] + rates[n // 2]) / 2
        out.append(e)
    return out


def from_job_tabs(resolver=None):
    """Crew from job workbooks embedded as rolodex tabs.

    ShowPhaze-era rows carry positions only: no rates (AKAV didn't run
    payroll then) and no grades (none were recorded).
    """
    resolver = resolver or RoleResolver()
    out = []
    for s in _spec_rows():
        if s["kind"] != "show":
            continue
        path = os.path.join(config.rollies_dir(), s["workbook"])
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        job, recs = parse_jobtab.parse_tab(wb[s["tab"]], s["tab"], resolver)
        wb.close()
        if not job:
            continue
        for r in recs:
            out.append({
                "source": "jobtab",
                "name": r["name"], "aka": r["aka"],
                "email": r["email"], "phoneDigits": r["phoneDigits"],
                "jobTitles": r["positions"],
                "source_file": s["workbook"], "source_tab": s["tab"],
                "jobLabel": job["label"],
            })
    return out


def from_roster(endpoint, token):
    """The people already on the Sheet -- as an IDENTITY source.

    A roster row carrying one person's email and another's name is the
    Sheet asserting they are the same human ('Tony Garcia' with
    anthony.garcia@..., 'Ben Hong' with benjamin.hong@...). Feeding those
    rows into the merge lets union-find collapse our split records the same
    way, so an upload can't have two of our people writing to one row and
    the second silently overwriting the first.

    It carries no attributes -- only identity. Everything else still comes
    from the source files.
    """
    from . import uploader
    rows = uploader.fetch_roster(endpoint, token) or []
    return [{
        "source": "roster",
        "name": r.get("name", ""),
        "email": r.get("email", ""),
        "phoneDigits": r.get("phoneDigits", ""),
        "sheetRow": r.get("row"),
    } for r in rows]


def collect_all(resolver=None, log=print, endpoint=None, token=None):
    """Every source, ready for merge.build_people()."""
    resolver = resolver or RoleResolver()
    recs, counts = [], {}

    rl = from_rollies(resolver)
    recs += rl; counts["rolly"] = len(rl); log("  rollies      %5d rows" % len(rl))

    vc = from_vcard()
    recs += vc; counts["vcard"] = len(vc); log("  contact list %5d do-not-hire" % len(vc))

    vd, vstats = from_vendors()
    recs += vd; counts["vendor"] = len(vd); log("  vendor list  %5d people" % len(vd))

    ct = from_contracts(resolver)
    recs += ct; counts["contract"] = len(ct); log("  contracts    %5d people" % len(ct))

    jt = from_job_tabs(resolver)
    recs += jt; counts["jobtab"] = len(jt); log("  job tabs     %5d crew rows" % len(jt))

    if endpoint and token:
        rs = from_roster(endpoint, token)
        recs += rs; counts["roster"] = len(rs)
        log("  live sheet   %5d rows (identity only)" % len(rs))

    return recs, counts
